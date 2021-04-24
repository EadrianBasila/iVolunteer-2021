import cv2
import numpy as np
from pyzbar.pyzbar import decode
import qrcode   
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image  
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

wb = openpyxl.load_workbook("dataSource\dummyData-iVolunteer.xlsx")
sheets = wb.sheetnames
sh1 = wb['Sheet1']

cap = cv2.VideoCapture(0)
cap.set(3,250)
cap.set(4,250)
row = sh1.max_row
column = sh1.max_column 

attendanceGreen = PatternFill(start_color='00ff77',
                   end_color='00ff77',
                   fill_type='solid')

idList = []
attendanceIn = []

for i in range(2,row+1): 
    qrVal = sh1.cell(i,1).value
    #print("ID val: "+qrVal,"added.")
    idList.append(qrVal)
print("ID Loaded.\n")

def attendanceCheck(qrData):
    A = 1
    i = 2
    dataVal = str(qrData)
    while A!=0:
        refVal = str(sh1.cell(i,1).value)
        if dataVal == refVal:
            anchorFill = "G"+str(i)
            print("Welcome", str(sh1.cell(i,2).value))
            sh1[str(anchorFill)].fill = attendanceGreen
            attendanceIn.append(qrData)
            wb.save("dataSource\dummyData-iVolunteer.xlsx") 
            A = 0

        else:
            i += 1
           
while True:
    success, img = cap.read()

    for qrcode in decode(img):
        qrData = qrcode.data.decode('utf-8')
        #print (qrData)
        
        if qrData in idList:
            if qrData in attendanceIn:
                qrStatus = 'Registered'
                polyColor = (0,229,255)             
            else:     
                qrStatus = 'Registering'
                polyColor = (0,255,0)           
                attendanceCheck(qrData)          
        else:
            qrStatus = 'Not Registered'
            polyColor = (0,0,255)
            
        polyPoints = np.array([qrcode.polygon],np.int32)
        polyPoints= polyPoints.reshape((-1,1,2))
        cv2.polylines(img,[polyPoints],True,polyColor,5)
        pts2 = qrcode.rect
        cv2.putText(img,qrStatus,(pts2[0],pts2[1]),cv2.FONT_HERSHEY_SIMPLEX, 0.9,polyColor,2)
 
    cv2.imshow('iVolunteer ScanQR', img)
    cv2.waitKey(1)