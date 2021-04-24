import qrcode   
import openpyxl
import numpy as np
from openpyxl.drawing.image import Image  
from openpyxl import Workbook
from openpyxl.drawing.image import Image  
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

wb = openpyxl.load_workbook("dataSource\dummyData-iVolunteer.xlsx")
sheets = wb.sheetnames
sh1 = wb['Sheet1']

qr = qrcode.QRCode(version=1, 
                   error_correction = qrcode.constants.ERROR_CORRECT_L, 
                   box_size = 20, border = 2)

attendanceRed = PatternFill(start_color='FF0000',
                   end_color='FF0000',
                   fill_type='solid')

def valtoQR(qrVal,qrName,anchorVal,i):
    imgName = str(qrName)
    imgID = str(qrVal)
    img = qrcode.make(imgID)
    qrLoc = str("imageBox//"+imgName+".png")
    img.save(qrLoc)
    qrToExcel(qrLoc,anchorVal,i)

#val to qr
def qrToExcel(qrLoc,anchorVal,i):
    imgAnchor = anchorVal
    imgA = Image(qrLoc)
    imgA.height = 50
    imgA.width = 50
    imgA.anchor = imgAnchor
    sh1.add_image(imgA)
    anchorFill = "G"+str(i)
    sh1[str(anchorFill)].fill = attendanceRed
    ################ PUSH EMAIL WITH ATTACHMENTS #######################
    ################ CALL iVolunteer_PushNoti.py #######################
    wb.save("dataSource\dummyData-iVolunteer.xlsx")

row = sh1.max_row
column = sh1.max_column 

#id to qr

for i in range(2,row+1): 
    qrVal = sh1.cell(i,1).value
    qrName = sh1.cell(i,2).value
    anchorVal = "F"+str(i)
    print("Object Loaded: ID[{0}] [{1}] [{2}]".format(qrVal,qrName,anchorVal))
    valtoQR(qrVal,qrName,anchorVal,i)
print("QR image loaded.")





